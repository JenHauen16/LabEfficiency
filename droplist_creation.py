#parses information from an excel file of daily patient information to create
#a list of patients and the cultures to drop slides on
#use the DLmarco.xlsm to format the drop list

import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import os
import re
import openpyxl
import datetime
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
wsname = str(date.today()- timedelta(days=1))
ws = wb.create_sheet(index=0, title=wsname)
wb.create_sheet(index = 1, title="Completed Cases")

ws['A1'] = "Chromosome_Cases_Set_" + str(date.today()- timedelta(days=1))
ws['A2'] = "Path_ID"
ws['B2'] = "Patient_Name"
ws['C2'] = "Cultures/Specimen Type"
ws['E1'] = 'Temp/Humidity:'
ws['E2'] = 'Slides_Made_and_Quality'

ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 15
ws.column_dimensions['J'].width = 15
ws.column_dimensions['K'].width = 15
ws.column_dimensions['L'].width = 15
ws.column_dimensions['M'].width = 15
ws.column_dimensions['N'].width = 15
ws.column_dimensions['O'].width = 15
ws.column_dimensions['P'].width = 15
ws.column_dimensions['Q'].width = 15
ws.column_dimensions['R'].width = 15
ws.column_dimensions['S'].width = 15
ws.column_dimensions['T'].width = 15
ws.column_dimensions['U'].width = 20


#os.chdir('C:\\Users\\CytoLab\\Desktop\\Shared drive docs')
file = "Daily Log07.31.19.xlsx"
#file = 'Daily Log' + date.today().strftime("%m.%d.%y") + '.xlsx'
x1 = pd.ExcelFile(file)
print(x1.sheet_names)
df1 = x1.parse('Case List')


for index, row in df1.iterrows():
    new_line_num = ws.max_row + 5
    if 'Chromosome' in row[4]:
        print(row[0], row[1], row[7][0:5], row[7][12:18])
        ws.cell(column=1, row=new_line_num, value=row[0])
        ws.cell(column=2, row=new_line_num, value=row[1])
        ws.cell(column=5, row=new_line_num, value=row[7])
        ws.cell(column=3, row=new_line_num, value=row[12])
    else:
        continue

dv1 = DataValidation(type="list", formula1='"8-6, 8-7, 8-8, 8-9, 8-10"', allow_blank=True)
ws.add_data_validation(dv1)
dv1.add('E8:K8')

dv2 = DataValidation(type="list", formula1='"CR, PLF, RM, PAE, CO, JR, TS, CG, Nella, DC, BKM, JH, YA, AY"', allow_blank=True)
ws.add_data_validation(dv2)
dv2.add("E9:K9")

dv3 = DataValidation(type="list", formula1='"very poor, poor, fair, good, very good, no mets"', allow_blank=True)
ws.add_data_validation(dv3)
dv3.add("E10:K10")

ws['D8'] = 'Date:'
ws['D9'] = 'Tech:'
ws['D10'] = 'Quality:'
ws['D11'] = 'Comments:'

wb.remove(wb['Sheet'])

wb.save('ChromosomeCases Week of' + '_' + str(date.today() - timedelta(days=1)) + '.xlsx')                
wb.close()



