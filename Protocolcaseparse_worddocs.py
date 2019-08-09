#parses "protocol" cases from multiple Word documents and puts in one Excel document

import os
import openpyxl
import datetime
from datetime import date
from openpyxl import Workbook
import docx
from docx import Document

wb = Workbook()
ws = wb.active
ws['A1'] = "Protocol Cases As of " + str(date.today())
ws['A2'] = "Patient"
ws['B2'] = 'Path ID'
ws['C2'] = 'Cultures Set'

os.chdir('C:\\Users\\CytoLab\\AppData\\Local\\Programs\\Python\\Python37-32\\DROP LIST')
for dirpath, dirnames, filenames in os.walk('C:\\Users\\CytoLab\\AppData\\Local\\Programs\\Python\\Python37-32\\DROP LIST'):
    for file in filenames:
        j = os.path.join(dirpath,file)
        doc = Document(j)
        tables = doc.tables
        for table in tables:
            for row in table.rows:
                new_line_num = ws.max_row + 1
                if "PROTOCOL" in row.cells[2].text:
                    ws.cell(column=1, row=new_line_num, value=row.cells[1].text)
                    ws.cell(column=2, row=new_line_num, value=row.cells[0].text)
                    ws.cell(column=3, row=new_line_num, value=row.cells[2].text)
                else:
                    continue
wb.save('protcolcases' + '_' + str(date.today()) + '.xlsx')                
wb.close()
